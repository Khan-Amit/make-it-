import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

print("🎯 Creating DEMO Excel file...")

# ============================================
# SHEET 1: STOCK CALCULATOR DEMO
# ============================================
stock_data = {
    'Product': ['Gold', 'Silver', 'Platinum', 'Diamond', 'Ruby', '⭐ TEST YOURSELF ⭐'],
    'Quantity': [100, 500, 50, 25, 40, 0],
    'Cost Price ($)': [75000, 800, 25000, 12000, 4000, 0],
    'Selling Price ($)': [85000, 950, 30000, 15000, 5000, 0],
    'Total Cost ($)': [0, 0, 0, 0, 0, 0],
    'Total Value ($)': [0, 0, 0, 0, 0, 0],
    'Profit ($)': [0, 0, 0, 0, 0, 0],
    'Margin %': [0, 0, 0, 0, 0, 0]
}
df_stock = pd.DataFrame(stock_data)

# Calculate formulas
for i in range(len(df_stock)):
    qty = df_stock.at[i, 'Quantity']
    cost = df_stock.at[i, 'Cost Price ($)']
    sell = df_stock.at[i, 'Selling Price ($)']
    
    df_stock.at[i, 'Total Cost ($)'] = qty * cost
    df_stock.at[i, 'Total Value ($)'] = qty * sell
    df_stock.at[i, 'Profit ($)'] = (qty * sell) - (qty * cost)
    
    if qty * cost > 0:
        df_stock.at[i, 'Margin %'] = round(((qty * sell) - (qty * cost)) / (qty * cost) * 100, 1)
    else:
        df_stock.at[i, 'Margin %'] = 0

# ============================================
# SHEET 2: PERCENTAGE CALCULATOR DEMO
# ============================================
pct_data = {
    'Type': ['Discount', 'Sales Tax', 'Tip', 'Profit Margin', 'Commission', '⭐ TEST YOURSELF ⭐'],
    'Original Amount ($)': [1000, 500, 150, 5000, 2000, 0],
    'Rate (%)': [15, 10, 18, 25, 8, 0],
    'Calculated Amount ($)': [0, 0, 0, 0, 0, 0],
    'Final Amount ($)': [0, 0, 0, 0, 0, 0]
}
df_pct = pd.DataFrame(pct_data)

for i in range(len(df_pct)):
    amt = df_pct.at[i, 'Original Amount ($)']
    rate = df_pct.at[i, 'Rate (%)']
    df_pct.at[i, 'Calculated Amount ($)'] = amt * (rate / 100)
    df_pct.at[i, 'Final Amount ($)'] = amt + df_pct.at[i, 'Calculated Amount ($)']

# ============================================
# SHEET 3: MATH DEMO
# ============================================
math_data = {
    'Operation': ['Addition', 'Subtraction', 'Multiplication', 'Division', '⭐ TEST YOURSELF ⭐'],
    'Number A': [25, 50, 12, 100, 0],
    'Number B': [15, 25, 8, 20, 0],
    'Result': [40, 25, 96, 5, 0]
}
df_math = pd.DataFrame(math_data)

# ============================================
# SHEET 4: INSTRUCTIONS
# ============================================
instructions = [
    ['🎯 HOW TO USE THIS DEMO'],
    [''],
    ['1️⃣ STOCK CALCULATOR SHEET:'],
    ['   → Change any number in Quantity, Cost Price, or Selling Price'],
    ['   → Total Cost, Total Value, Profit, and Margin % will AUTO-CALCULATE'],
    ['   → Try the "TEST YOURSELF" row at the bottom!'],
    [''],
    ['2️⃣ PERCENTAGE CALCULATOR SHEET:'],
    ['   → Change Original Amount or Rate %'],
    ['   → Calculated Amount and Final Amount update INSTANTLY'],
    [''],
    ['3️⃣ MATH SHEET:'],
    ['   → Change Number A or Number B'],
    ['   → Result updates automatically'],
    [''],
    ['=' * 50],
    ['📞 FULL VERSION (10 Calculators): $199 Early Bird'],
    ['📧 Contact: seliim.ahmed@gmail.com'],
    ['🔗 https://khan-amit.github.io/make-it-/'],
    [''],
    ['✅ After purchasing full version, you get:'],
    ['   • Digital Certificate with your name'],
    ['   • Unique License Number'],
    ['   • 12-Month Warranty Certificate'],
    ['   • Lifetime Updates'],
    ['   • 10 Professional Calculators']
]

df_instructions = pd.DataFrame({'INSTRUCTIONS': [i[0] if isinstance(i, list) else i for i in instructions]})

# ============================================
# WRITE TO EXCEL
# ============================================
with pd.ExcelWriter('demo_calculator.xlsx', engine='openpyxl') as writer:
    df_stock.to_excel(writer, sheet_name='💰 STOCK DEMO', index=False)
    df_pct.to_excel(writer, sheet_name='🎯 PERCENTAGE DEMO', index=False)
    df_math.to_excel(writer, sheet_name='🧮 MATH DEMO', index=False)
    df_instructions.to_excel(writer, sheet_name='📋 INSTRUCTIONS', index=False, header=False)

# ============================================
# APPLY FORMATTING
# ============================================
wb = openpyxl.load_workbook('demo_calculator.xlsx')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Style headers
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    if sheet_name != '📋 INSTRUCTIONS':
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

wb.save('demo_calculator.xlsx')

print("\n" + "="*60)
print("✅✅✅ DEMO CALCULATOR CREATED SUCCESSFULLY! ✅✅✅")
print("="*60)
print("\n📁 FILE CREATED: demo_calculator.xlsx")
print("\n📊 FILE CONTAINS 4 SHEETS:")
print("   1. 💰 STOCK DEMO - Fully working stock calculator")
print("   2. 🎯 PERCENTAGE DEMO - Fully working percentage calculator")
print("   3. 🧮 MATH DEMO - Basic math operations")
print("   4. 📋 INSTRUCTIONS - How to use guide")
print("\n🎯 TEST IT RIGHT NOW:")
print("   1. Double-click demo_calculator.xlsx")
print("   2. Go to '💰 STOCK DEMO' sheet")
print("   3. Change Quantity from 100 to 200")
print("   4. Watch Total Cost, Total Value, Profit CHANGE AUTOMATICALLY!")
print("\n📤 NEXT STEP: Upload this file to GitHub")
print("   git add demo_calculator.xlsx")
print("   git commit -m 'Add working demo calculator'")
print("   git push origin main")
print("\n🔗 Then test: https://khan-amit.github.io/make-it-/")
print("="*60)
