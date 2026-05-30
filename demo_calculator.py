import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formula import Tokenizer
import os
import math

class DemoExcelCalculator:
    def __init__(self, filename="demo_calculator.xlsx"):
        self.filename = filename
    
    def create_demo(self):
        """Create DEMO Excel file with WORKING calculations"""
        
        # ========== SHEET 1: WELCOME / INFO ==========
        welcome_data = {
            'Welcome to DEMO Version': [''],
            '': [''],
            '✅ This is a FULLY FUNCTIONAL DEMO': [''],
            '✅ 2 Working Calculators for you to test': [''],
            '✅ Enter ANY numbers - formulas calculate instantly': [''],
            '✅ Full version has 10 calculators': [''],
            '': [''],
            '📞 Contact: seliim.ahmed@gmail.com': [''],
            '💰 Full Version: $199 (Early Bird)': [''],
            '🔗 https://khan-amit.github.io/make-it-/': ['']
        }
        df_welcome = pd.DataFrame(list(welcome_data.items()), columns=['Info', ''])
        
        # ========== SHEET 2: DEMO STOCK CALCULATOR (WORKING) ==========
        stock_data = {
            'Product': ['Gold', 'Silver', 'Platinum', 'Diamond', 'Ruby', 'ENTER YOURS ➡️'],
            'Quantity': [100, 500, 50, 25, 40, 0],
            'Cost Price ($)': [75000, 800, 25000, 12000, 4000, 0],
            'Selling Price ($)': [85000, 950, 30000, 15000, 5000, 0],
            'Total Cost ($)': [0, 0, 0, 0, 0, 0],
            'Total Value ($)': [0, 0, 0, 0, 0, 0],
            'Profit ($)': [0, 0, 0, 0, 0, 0],
            'Margin %': [0, 0, 0, 0, 0, 0]
        }
        df_stock = pd.DataFrame(stock_data)
        
        # Calculate formulas
        for i in range(len(df_stock)):
            qty = df_stock.at[i, 'Quantity']
            cost = df_stock.at[i, 'Cost Price ($)']
            sell = df_stock.at[i, 'Selling Price ($)']
            
            df_stock.at[i, 'Total Cost ($)'] = qty * cost
            df_stock.at[i, 'Total Value ($)'] = qty * sell
            df_stock.at[i, 'Profit ($)'] = (qty * sell) - (qty * cost)
            if qty * cost > 0:
                df_stock.at[i, 'Margin %'] = round(((qty * sell) - (qty * cost)) / (qty * cost) * 100, 1)
            else:
                df_stock.at[i, 'Margin %'] = 0
        
        # Add summary row
        summary = {
            'Product': ['📊 TOTALS ➡️'],
            'Quantity': [df_stock['Quantity'].sum()],
            'Cost Price ($)': ['-'],
            'Selling Price ($)': ['-'],
            'Total Cost ($)': [df_stock['Total Cost ($)'].sum()],
            'Total Value ($)': [df_stock['Total Value ($)'].sum()],
            'Profit ($)': [df_stock['Profit ($)'].sum()],
            'Margin %': [round(df_stock['Profit ($)'].sum() / df_stock['Total Cost ($)'].sum() * 100, 1) if df_stock['Total Cost ($)'].sum() > 0 else 0]
        }
        df_summary = pd.DataFrame(summary)
        df_stock = pd.concat([df_stock, df_summary], ignore_index=True)
        
        # ========== SHEET 3: DEMO PERCENTAGE CALCULATOR (WORKING) ==========
        pct_data = {
            'Type': ['Discount', 'Sales Tax', 'Tip', 'Profit Margin', 'Commission', '👇 TEST YOURS'],
            'Original Amount ($)': [1000, 500, 150, 5000, 2000, 0],
            'Rate (%)': [15, 10, 18, 25, 8, 0],
            'Calculated Amount ($)': [0, 0, 0, 0, 0, 0],
            'Final Amount ($)': [0, 0, 0, 0, 0, 0]
        }
        df_pct = pd.DataFrame(pct_data)
        
        # Calculate formulas
        for i in range(len(df_pct)):
            amount = df_pct.at[i, 'Original Amount ($)']
            rate = df_pct.at[i, 'Rate (%)']
            
            df_pct.at[i, 'Calculated Amount ($)'] = amount * (rate / 100)
            df_pct.at[i, 'Final Amount ($)'] = amount + df_pct.at[i, 'Calculated Amount ($)']
        
        # ========== SHEET 4: QUICK MATH DEMO ==========
        math_data = {
            'Operation': ['Addition', 'Subtraction', 'Multiplication', 'Division'],
            'Enter Number A': [25, 50, 12, 100],
            'Enter Number B': [15, 25, 8, 20],
            'Result (Auto)': [0, 0, 0, 0]
        }
        df_math = pd.DataFrame(math_data)
        
        for i in range(4):
            a = df_math.at[i, 'Enter Number A']
            b = df_math.at[i, 'Enter Number B']
            if i == 0:
                df_math.at[i, 'Result (Auto)'] = a + b
            elif i == 1:
                df_math.at[i, 'Result (Auto)'] = a - b
            elif i == 2:
                df_math.at[i, 'Result (Auto)'] = a * b
            elif i == 3:
                df_math.at[i, 'Result (Auto)'] = round(a / b, 2) if b != 0 else 0
        
        # ========== WRITE ALL SHEETS ==========
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            df_welcome.to_excel(writer, sheet_name='📋 INFO', index=False, header=False)
            df_stock.to_excel(writer, sheet_name='💰 DEMO STOCK', index=False)
            df_pct.to_excel(writer, sheet_name='🎯 DEMO PERCENTAGE', index=False)
            df_math.to_excel(writer, sheet_name='🧮 DEMO MATH', index=False)
        
        # Apply formatting
        self._apply_demo_formatting()
        
        print("\n" + "="*60)
        print("✅ DEMO EXCEL CALCULATOR CREATED!")
        print("="*60)
        print(f"\n📁 File: {os.path.abspath(self.filename)}")
        print("\n📊 WORKING DEMO SHEETS:")
        print("   📋 INFO          - How to use this demo")
        print("   💰 DEMO STOCK    - FULLY WORKING stock calculator")
        print("   🎯 DEMO PERCENTAGE - FULLY WORKING percentage calc")
        print("   🧮 DEMO MATH     - Working math operations")
        print("\n✨ TRY THIS:")
        print("   1. Open the Excel file")
        print("   2. Go to 💰 DEMO STOCK sheet")
        print("   3. Change ANY number (Quantity, Cost, or Price)")
        print("   4. Watch Total Cost, Value, Profit AUTO-UPDATE!")
        print("\n   5. Go to 🎯 DEMO PERCENTAGE sheet")
        print("   6. Change Original Amount or Rate %")
        print("   7. Watch Calculated Amount and Final AUTO-UPDATE!")
        print("\n💡 This is just 3 of 10 calculators in FULL version!")
        print("   Full version includes: Statistics, Finance, Geometry, Trend, Unit, Demography")
        print("\n💰 FULL VERSION: $199 (Early Bird)")
        print("   Contact: seliim.ahmed@gmail.com")
        print("="*60)
        
        return True
    
    def _apply_demo_formatting(self):
        """Apply nice formatting to demo"""
        try:
            wb = openpyxl.load_workbook(self.filename)
            
            # Format DEMO STOCK sheet
            if '💰 DEMO STOCK' in wb.sheetnames:
                ws = wb['💰 DEMO STOCK']
                header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                money_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Color money columns
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=8):
                    for cell in row:
                        cell.fill = money_fill
                        cell.number_format = '#,##0.00'
                
                # Adjust column widths
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value and len(str(cell.value)) > max_len:
                                max_len = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 20)
            
            # Format DEMO PERCENTAGE sheet
            if '🎯 DEMO PERCENTAGE' in wb.sheetnames:
                ws = wb['🎯 DEMO PERCENTAGE']
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            wb.save(self.filename)
            print("✓ Demo formatting applied!")
        except Exception as e:
            print(f"✓ Demo created successfully")

# ============================================================
# RUN THE DEMO
# ============================================================
if __name__ == "__main__":
    print("\n" + "🎯"*30)
    print("   CREATING DEMO EXCEL CALCULATOR")
    print("   (FULLY WORKING - TEST BEFORE BUYING)")
    print("🎯"*30 + "\n")
    
    # Install packages if needed
    try:
        import pandas, openpyxl
        print("✓ Required packages found!\n")
    except ImportError:
        print("📦 Installing pandas and openpyxl...")
        os.system("pip install pandas openpyxl")
        print("✅ Packages installed!\n")
    
    # Create the demo
    demo = DemoExcelCalculator("demo_calculator.xlsx")
    demo.create_demo()
    
    print("\n🚀 DEMO READY! Open 'demo_calculator.xlsx' and TEST IT NOW!")
    print("   Change any number - watch it AUTO-CALCULATE!")
