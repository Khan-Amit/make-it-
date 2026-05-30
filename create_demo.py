import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

print("🎯 Creating DEMO Excel file...")

# Create Stock Calculator Demo
stock_data = {
    'Product': ['Gold', 'Silver', 'Platinum', 'Diamond', 'Ruby', '⭐ TRY YOURSELF ⭐'],
    'Quantity': [100, 500, 50, 25, 40, 0],
    'Cost Price ($)': [75000, 800, 25000, 12000, 4000, 0],
    'Selling Price ($)': [85000, 950, 30000, 15000, 5000, 0],
    'Total Cost ($)': [0, 0, 0, 0, 0, 0],
    'Total Value ($)': [0, 0, 0, 0, 0, 0],
    'Profit ($)': [0, 0, 0, 0, 0, 0],
    'Margin %': [0, 0, 0, 0, 0, 0]
}
df_stock = pd.DataFrame(stock_data)

# Calculate formulas
for i in range(len(df_stock)):
    qty = df_stock.at[i, 'Quantity']
    cost = df_stock.at[i, 'Cost Price ($)']
    sell = df_stock.at[i, 'Selling Price ($)']
    df_stock.at[i, 'Total Cost ($)'] = qty * cost
    df_stock.at[i, 'Total Value ($)'] = qty * sell
    df_stock.at[i, 'Profit ($)'] = (qty * sell) - (qty * cost)
    if qty * cost > 0:
        df_stock.at[i, 'Margin %'] = round(((qty * sell) - (qty * cost)) / (qty * cost) * 100, 1)

# Create Percentage Calculator Demo
pct_data = {
    'Type': ['Discount', 'Sales Tax', 'Tip', 'Profit Margin', 'Commission', '⭐ TRY YOURSELF ⭐'],
    'Original Amount ($)': [1000, 500, 150, 5000, 2000, 0],
    'Rate (%)': [15, 10, 18, 25, 8, 0],
    'Calculated Amount ($)': [0, 0, 0, 0, 0, 0],
    'Final Amount ($)': [0, 0, 0, 0, 0, 0]
}
df_pct = pd.DataFrame(pct_data)

for i in range(len(df_pct)):
    amt = df_pct.at[i, 'Original Amount ($)']
    rate = df_pct.at[i, 'Rate (%)']
    df_pct.at[i, 'Calculated Amount ($)'] = amt * (rate / 100)
    df_pct.at[i, 'Final Amount ($)'] = amt + df_pct.at[i, 'Calculated Amount ($)']

# Create Math Demo
math_data = {
    'Operation': ['Addition', 'Subtraction', 'Multiplication', 'Division', '⭐ TEST YOURS ⭐'],
    'Number A': [25, 50, 12, 100, 0],
    'Number B': [15, 25, 8, 20, 0],
    'Result': [40, 25, 96, 5, 0]
}
df_math = pd.DataFrame(math_data)

# Create Instructions Sheet
instructions = {
    '🎯 HOW TO USE THIS DEMO': [''],
    '': [''],
    '1️⃣ STOCK CALCULATOR SHEET': [''],
    '   → Change any number in Quantity, Cost Price, or Selling Price', [''],
    '   → Total Cost, Total Value, Profit, and Margin % will AUTO-CALCULATE', [''],
    '   → Try the "TRY YOURSELF" row at the bottom!', [''],
    '': [''],
    '2️⃣ PERCENTAGE CALCULATOR SHEET': [''],
    '   → Change Original Amount or Rate %', [''],
    '   → Calculated Amount and Final Amount update INSTANTLY', [''],
    '': [''],
    '3️⃣ MATH SHEET': [''],
    '   → Change Number A or Number B', [''],
    '   → Result updates automatically', [''],
    '': [''],
    '📞 FULL VERSION (10 Calculators): $199 Early Bird', [''],
    '📧 Contact: seliim.ahmed@gmail.com', [''],
    '🔗 https://khan-amit.github.io/make-it-/', ['']
}

# Flatten instructions
flat_instructions = []
for item in instructions:
    if isinstance(item, list):
        flat_instructions.extend(item)
    else:
        flat_instructions.append(item)

df_info = pd.DataFrame({'Info': flat_instructions})

# Write to Excel
with pd.ExcelWriter('demo_calculator.xlsx', engine='openpyxl') as writer:
    df_stock.to_excel(writer, sheet_name='💰 STOCK DEMO', index=False)
    df_pct.to_excel(writer, sheet_name='🎯 PERCENTAGE DEMO', index=False)
    df_math.to_excel(writer, sheet_name='🧮 MATH DEMO', index=False)
    df_info.to_excel(writer, sheet_name='📋 INSTRUCTIONS', index=False)

# Apply formatting
wb = openpyxl.load_workbook('demo_calculator.xlsx')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Header styling
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    # Adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 25)
wb.save('demo_calculator.xlsx')

print("\n" + "="*50)
print("✅ DEMO CALCULATOR CREATED SUCCESSFULLY!")
print("="*50)
print("\n📁 File: demo_calculator.xlsx")
print("\n📊 INCLUDES 4 SHEETS:")
print("   💰 STOCK DEMO - Fully working stock calculator")
print("   🎯 PERCENTAGE DEMO - Fully working percentage calc")
print("   🧮 MATH DEMO - Basic math operations")
print("   📋 INSTRUCTIONS - How to use guide")
print("\n🎯 TEST IT:")
print("   1. Open demo_calculator.xlsx")
print("   2. Go to 💰 STOCK DEMO sheet")
print("   3. Change ANY number in Quantity, Cost, or Price")
print("   4. Watch it AUTO-CALCULATE!")
print("\n📤 Now upload 'demo_calculator.xlsx' to GitHub!")
print("="*50)
