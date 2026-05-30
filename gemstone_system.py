import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime
import os

class GemstoneStockSystem:
    def __init__(self, filename="gemstone_stock.xlsx"):
        self.filename = filename
        self.current_stock = None
        
    def create_gemstone_system(self):
        """Create complete gemstone stock management system"""
        
        # 1. STOCK SHEET - Main inventory
        stock_data = {
            'Gemstone ID': ['G001', 'G002', 'G003', 'G004', 'G005', 'G006', 'G007', 'G008'],
            'Gemstone Name': ['Ruby', 'Emerald', 'Sapphire Blue', 'Sapphire Pink', 'Diamond', 'Amethyst', 'Topaz', 'Opal'],
            'Type': ['Precious', 'Precious', 'Precious', 'Precious', 'Precious', 'Semi-Precious', 'Semi-Precious', 'Semi-Precious'],
            'Carat Weight': [1.25, 1.50, 2.00, 1.75, 0.75, 3.00, 2.50, 1.80],
            'Quantity': [5, 3, 4, 2, 6, 10, 8, 4],
            'Cost per Carat ($)': [5000, 8000, 6000, 12000, 15000, 200, 150, 300],
            'Total Cost ($)': [0, 0, 0, 0, 0, 0, 0, 0],
            'Selling Price ($)': [7500, 12000, 9000, 18000, 22000, 350, 250, 500],
            'Total Value ($)': [0, 0, 0, 0, 0, 0, 0, 0],
            'Profit ($)': [0, 0, 0, 0, 0, 0, 0, 0],
            'Status': ['In Stock', 'In Stock', 'Low Stock', 'In Stock', 'Premium', 'In Stock', 'Low Stock', 'In Stock']
        }
        
        df_stock = pd.DataFrame(stock_data)
        
        # Calculate formulas
        df_stock['Total Cost ($)'] = df_stock['Carat Weight'] * df_stock['Cost per Carat ($)'] * df_stock['Quantity']
        df_stock['Total Value ($)'] = df_stock['Carat Weight'] * df_stock['Selling Price ($)'] * df_stock['Quantity']
        df_stock['Profit ($)'] = df_stock['Total Value ($)'] - df_stock['Total Cost ($)']
        
        # 2. CALCULATIONS SHEET - Summary metrics
        calculations_data = {
            'Metric': [
                'Total Inventory Value',
                'Total Investment Cost',
                'Total Potential Profit',
                'Average Profit Margin',
                'Total Carats in Stock',
                'Total Units',
                'Precious Stones Count',
                'Semi-Precious Stones Count',
                'Low Stock Items',
                'Most Valuable Stone',
                'Last Updated'
            ],
            'Value': [
                f"${df_stock['Total Value ($)'].sum():,.2f}",
                f"${df_stock['Total Cost ($)'].sum():,.2f}",
                f"${df_stock['Profit ($)'].sum():,.2f}",
                f"{(df_stock['Profit ($)'].sum() / df_stock['Total Cost ($)'].sum() * 100):.1f}%",
                f"{df_stock['Carat Weight'].sum():.2f}",
                df_stock['Quantity'].sum(),
                len(df_stock[df_stock['Type'] == 'Precious']),
                len(df_stock[df_stock['Type'] == 'Semi-Precious']),
                len(df_stock[df_stock['Status'] == 'Low Stock']),
                df_stock.loc[df_stock['Total Value ($)'].idxmax(), 'Gemstone Name'],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
        }
        df_calculations = pd.DataFrame(calculations_data)
        
        # 3. TRANSACTIONS SHEET - Sales record
        transactions_data = {
            'Date': ['2026-05-15', '2026-05-20', '2026-05-25', '2026-05-28'],
            'Gemstone': ['Ruby', 'Emerald', 'Diamond', 'Sapphire Blue'],
            'Carats Sold': [0.50, 0.75, 0.25, 1.00],
            'Quantity Sold': [1, 1, 1, 1],
            'Sale Price ($)': [8000, 13000, 25000, 10000],
            'Revenue ($)': [8000, 13000, 25000, 10000],
            'Profit ($)': [3000, 4500, 10000, 4000]
        }
        df_transactions = pd.DataFrame(transactions_data)
        
        # 4. PRICE CALCULATOR SHEET
        price_data = {
            'Gemstone': ['Ruby', 'Emerald', 'Sapphire', 'Diamond', 'Amethyst'],
            'Base Price/Carat ($)': [5000, 8000, 6000, 15000, 200],
            'Quality Multiplier': [1.0, 1.0, 1.0, 1.0, 1.0],
            'Final Price/Carat ($)': [0, 0, 0, 0, 0],
            'Calculator Formula': ['=Carat_Weight * Final_Price', '=Carat_Weight * Final_Price', '=Carat_Weight * Final_Price', '=Carat_Weight * Final_Price', '=Carat_Weight * Final_Price']
        }
        df_price = pd.DataFrame(price_data)
        df_price['Final Price/Carat ($)'] = df_price['Base Price/Carat ($)'] * df_price['Quality Multiplier']
        
        # Write to Excel with formatting
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            df_stock.to_excel(writer, sheet_name='STOCK_INVENTORY', index=False)
            df_calculations.to_excel(writer, sheet_name='CALCULATIONS', index=False)
            df_transactions.to_excel(writer, sheet_name='TRANSACTIONS', index=False)
            df_price.to_excel(writer, sheet_name='PRICE_CALCULATOR', index=False)
            
            # Add empty sales input sheet
            empty_sales = pd.DataFrame(columns=['Date', 'Gemstone ID', 'Quantity Sold', 'Selling Price'])
            empty_sales.to_excel(writer, sheet_name='SALES_INPUT', index=False)
        
        # Apply formatting
        self._format_workbook()
        
        print(f"✓ Gemstone Stock System created: {self.filename}")
        self._display_summary(df_stock, df_calculations)
        
        return True
    
    def _format_workbook(self):
        """Apply professional formatting to all sheets"""
        try:
            wb = openpyxl.load_workbook(self.filename)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            money_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format headers
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Color money columns
                if sheet_name == "STOCK_INVENTORY":
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=10):
                        for cell in row:
                            cell.fill = money_fill
                            cell.number_format = '$#,##0.00'
            
            wb.save(self.filename)
            print("✓ Formatting applied")
            
        except Exception as e:
            print(f"Formatting note: {e}")
    
    def _display_summary(self, df_stock, df_calculations):
        """Display stock summary"""
        print("\n" + "="*50)
        print("📊 GEMSTONE STOCK SUMMARY")
        print("="*50)
        print(f"💎 Total Inventory Value: ${df_stock['Total Value ($)'].sum():,.2f}")
        print(f"💰 Total Investment: ${df_stock['Total Cost ($)'].sum():,.2f}")
        print(f"📈 Potential Profit: ${df_stock['Profit ($)'].sum():,.2f}")
        print(f"📊 Profit Margin: {(df_stock['Profit ($)'].sum() / df_stock['Total Cost ($)'].sum() * 100):.1f}%")
        print(f"⚖️ Total Carats: {df_stock['Carat Weight'].sum():.2f}")
        print(f"📦 Total Units: {df_stock['Quantity'].sum()}")
        print("="*50)
        
        # Top 3 valuable stones
        print("\n🏆 MOST VALUABLE GEMSTONES:")
        top3 = df_stock.nlargest(3, 'Total Value ($)')[['Gemstone Name', 'Total Value ($)']]
        for idx, row in top3.iterrows():
            print(f"   • {row['Gemstone Name']}: ${row['Total Value ($)']:,.2f}")
    
    def add_sale_transaction(self, gemstone_id, quantity_sold):
        """Record a sale and update inventory"""
        try:
            df_stock = pd.read_excel(self.filename, sheet_name='STOCK_INVENTORY')
            
            # Find gemstone
            mask = df_stock['Gemstone ID'] == gemstone_id
            if not mask.any():
                print(f"✗ Gemstone ID {gemstone_id} not found")
                return False
            
            current_qty = df_stock.loc[mask, 'Quantity'].values[0]
            if current_qty < quantity_sold:
                print(f"✗ Insufficient stock! Only {current_qty} available")
                return False
            
            # Update quantity
            df_stock.loc[mask, 'Quantity'] = current_qty - quantity_sold
            
            # Update status
            if df_stock.loc[mask, 'Quantity'].values[0] == 0:
                df_stock.loc[mask, 'Status'] = 'Out of Stock'
            elif df_stock.loc[mask, 'Quantity'].values[0] <= 2:
                df_stock.loc[mask, 'Status'] = 'Low Stock'
            
            # Recalculate totals
            df_stock['Total Cost ($)'] = df_stock['Carat Weight'] * df_stock['Cost per Carat ($)'] * df_stock['Quantity']
            df_stock['Total Value ($)'] = df_stock['Carat Weight'] * df_stock['Selling Price ($)'] * df_stock['Quantity']
            df_stock['Profit ($)'] = df_stock['Total Value ($)'] - df_stock['Total Cost ($)']
            
            # Save updated stock
            with pd.ExcelWriter(self.filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_stock.to_excel(writer, sheet_name='STOCK_INVENTORY', index=False)
            
            print(f"✓ Sold {quantity_sold} of {gemstone_id}")
            print(f"  Remaining: {df_stock.loc[mask, 'Quantity'].values[0]} units")
            return True
            
        except Exception as e:
            print(f"✗ Error recording sale: {e}")
            return False
    
    def calculate_price(self, gemstone, carat_weight, quality_multiplier=1.0):
        """Calculate selling price for a gemstone"""
        price_data = {
            'Ruby': 5000, 'Emerald': 8000, 'Sapphire': 6000,
            'Diamond': 15000, 'Amethyst': 200, 'Topaz': 150, 'Opal': 300
        }
        
        if gemstone not in price_data:
            print(f"✗ Gemstone '{gemstone}' not in database")
            return None
        
        base_price = price_data[gemstone]
        final_price = base_price * quality_multiplier
        total_value = final_price * carat_weight
        
        print(f"\n💎 PRICE CALCULATION:")
        print(f"   Gemstone: {gemstone}")
        print(f"   Carat Weight: {carat_weight}")
        print(f"   Quality Multiplier: {quality_multiplier}x")
        print(f"   Price per Carat: ${final_price:,.2f}")
        print(f"   TOTAL VALUE: ${total_value:,.2f}")
        
        return total_value

def main():
    print("=" * 55)
    print("💎 GEMSTONE STOCK CALCULATION SYSTEM")
    print("=" * 55)
    
    system = GemstoneStockSystem("gemstone_stock.xlsx")
    
    # Create the system
    system.create_gemstone_system()
    
    # Demo calculations
    print("\n🔮 TRY THE CALCULATOR:")
    system.calculate_price("Diamond", 1.5, 1.2)
    system.calculate_price("Ruby", 2.0, 1.0)
    
    # Demo sale
    print("\n📦 SALE DEMO:")
    system.add_sale_transaction("G001", 1)  # Sell 1 Ruby
    
    print("\n" + "=" * 55)
    print("✅ SYSTEM READY!")
    print(f"📁 File created: {os.path.abspath('gemstone_stock.xlsx')}")
    print("\n📌 FEATURES:")
    print("   • Auto-calculating inventory")
    print("   • Profit tracking")
    print("   • Stock alerts (Low Stock status)")
    print("   • Sales recording")
    print("   • Price calculator")
    print("=" * 55)

if __name__ == "__main__":
    try:
        import pandas, openpyxl
    except ImportError:
        print("Installing packages...")
        os.system("pip install pandas openpyxl")
    
    main()
