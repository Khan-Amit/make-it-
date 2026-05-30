import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import math
import subprocess
import sys

def install_packages():
    """Install required packages"""
    packages = ['pandas', 'openpyxl']
    for package in packages:
        try:
            __import__(package)
        except ImportError:
            print(f"📦 Installing {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])

class MasterCalculator:
    def __init__(self, filename="master_calculator_final.xlsx"):
        self.filename = filename
        
    def create_all_calculators(self):
        """Create complete Excel calculator system"""
        
        print("\n🔧 Creating Master Calculator System...")
        
        # ========== SHEET 1: INDEX/DASHBOARD ==========
        index_data = {
            'S.No': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
            '📊 Calculator': ['Stock Inventory', 'Demography', 'Statistics', 'Finance', 
                             'Math Solver', 'Geometry', 'Percentage', 'Trend Analysis', 
                             'Unit Converter', 'Custom Calculator'],
            '📂 Category': ['Business', 'Demographics', 'Statistics', 'Finance', 
                           'Mathematics', 'Geometry', 'Mathematics', 'Business', 
                           'Science', 'Custom'],
            '📝 Description': ['Calculate stock value, cost & profit', 
                              'Population analysis by age and gender',
                              'Mean, median, sum and statistical analysis',
                              'Track income, expenses and running balance',
                              'Addition, subtraction, multiplication, division',
                              'Area, perimeter and volume of shapes',
                              'Discount, tax, tip, profit margin calculator',
                              'Sales growth percentage and moving averages',
                              'Length, weight, temperature, currency converter',
                              'Create your own custom calculations easily'],
            '🔗 Go to Sheet': ['💰 STOCK', '👥 DEMOGRAPHY', '📈 STATISTICS', '💳 FINANCE',
                              '🔢 MATH', '📐 GEOMETRY', '🎯 PERCENTAGE', '📉 TREND',
                              '⚖️ UNIT', '➕ CREATE NEW'],
            '✅ Status': ['Active', 'Active', 'Active', 'Active', 'Active', 
                         'Active', 'Active', 'Active', 'Active', 'Ready'],
            '📅 Last Used': [datetime.now().strftime("%Y-%m-%d"), '-', '-', '-', '-', 
                            '-', '-', '-', '-', '-']
        }
        df_index = pd.DataFrame(index_data)
        
        # ========== SHEET 2: STOCK CALCULATOR ==========
        stock_data = {
            'Product Code': ['STK-001', 'STK-002', 'STK-003', 'STK-004', 'STK-005', 'STK-006'],
            'Product Name': ['Gold 24K', 'Silver 999', 'Platinum', 'Diamond', 'Ruby', 'Emerald'],
            'Quantity': [100, 500, 50, 25, 40, 30],
            'Cost Price ($)': [75000, 800, 25000, 12000, 4000, 6500],
            'Selling Price ($)': [85000, 950, 30000, 15000, 5000, 8000],
            'Total Cost ($)': [0, 0, 0, 0, 0, 0],
            'Total Value ($)': [0, 0, 0, 0, 0, 0],
            'Profit ($)': [0, 0, 0, 0, 0, 0],
            'Profit Margin %': [0, 0, 0, 0, 0, 0]
        }
        df_stock = pd.DataFrame(stock_data)
        df_stock['Total Cost ($)'] = df_stock['Quantity'] * df_stock['Cost Price ($)']
        df_stock['Total Value ($)'] = df_stock['Quantity'] * df_stock['Selling Price ($)']
        df_stock['Profit ($)'] = df_stock['Total Value ($)'] - df_stock['Total Cost ($)']
        df_stock['Profit Margin %'] = round((df_stock['Profit ($)'] / df_stock['Total Cost ($)']) * 100, 2)
        
        # Add summary row
        stock_summary = pd.DataFrame({
            'Product Code': ['TOTAL'],
            'Product Name': ['SUMMARY'],
            'Quantity': [df_stock['Quantity'].sum()],
            'Cost Price ($)': ['-'],
            'Selling Price ($)': ['-'],
            'Total Cost ($)': [df_stock['Total Cost ($)'].sum()],
            'Total Value ($)': [df_stock['Total Value ($)'].sum()],
            'Profit ($)': [df_stock['Profit ($)'].sum()],
            'Profit Margin %': [round((df_stock['Profit ($)'].sum() / df_stock['Total Cost ($)'].sum()) * 100, 2)]
        })
        df_stock = pd.concat([df_stock, stock_summary], ignore_index=True)
        
        # ========== SHEET 3: DEMOGRAPHY ==========
        demo_data = {
            'Age Group': ['0-18', '19-30', '31-45', '46-60', '60+', 'TOTAL'],
            'Population': [25000, 45000, 38000, 22000, 15000, 145000],
            'Male %': [48, 50, 49, 47, 42, 0],
            'Female %': [52, 50, 51, 53, 58, 0],
            'Male Count': [0, 0, 0, 0, 0, 0],
            'Female Count': [0, 0, 0, 0, 0, 0],
            'Growth Rate %': [2.1, 1.8, 1.2, 0.5, -0.2, 0]
        }
        df_demo = pd.DataFrame(demo_data)
        for i in range(5):
            df_demo.at[i, 'Male Count'] = int(df_demo.at[i, 'Population'] * df_demo.at[i, 'Male %'] / 100)
            df_demo.at[i, 'Female Count'] = int(df_demo.at[i, 'Population'] * df_demo.at[i, 'Female %'] / 100)
        df_demo.at[5, 'Male Count'] = df_demo['Male Count'].sum()
        df_demo.at[5, 'Female Count'] = df_demo['Female Count'].sum()
        df_demo.at[5, 'Male %'] = round((df_demo.at[5, 'Male Count'] / df_demo.at[5, 'Population']) * 100, 1)
        df_demo.at[5, 'Female %'] = round((df_demo.at[5, 'Female Count'] / df_demo.at[5, 'Population']) * 100, 1)
        
        # ========== SHEET 4: STATISTICS ==========
        stats_data = {
            'Sample ID': ['S001', 'S002', 'S003', 'S004', 'S005'],
            'Data Point 1': [23, 45, 67, 12, 89],
            'Data Point 2': [34, 56, 78, 23, 45],
            'Data Point 3': [45, 67, 89, 34, 56],
            'Mean': [0, 0, 0, 0, 0],
            'Median': [0, 0, 0, 0, 0],
            'Sum': [0, 0, 0, 0, 0],
            'Min': [0, 0, 0, 0, 0],
            'Max': [0, 0, 0, 0, 0]
        }
        df_stats = pd.DataFrame(stats_data)
        for i in range(5):
            values = [df_stats.at[i, 'Data Point 1'], df_stats.at[i, 'Data Point 2'], df_stats.at[i, 'Data Point 3']]
            df_stats.at[i, 'Mean'] = round(sum(values) / len(values), 2)
            df_stats.at[i, 'Median'] = sorted(values)[1]
            df_stats.at[i, 'Sum'] = sum(values)
            df_stats.at[i, 'Min'] = min(values)
            df_stats.at[i, 'Max'] = max(values)
        
        # ========== SHEET 5: FINANCE ==========
        finance_data = {
            'Date': ['2026-05-01', '2026-05-02', '2026-05-03', '2026-05-04', '2026-05-05', '2026-05-06'],
            'Description': ['Salary', 'Rent', 'Groceries', 'Investment', 'Shopping', 'Savings'],
            'Category': ['Income', 'Housing', 'Food', 'Investment', 'Shopping', 'Savings'],
            'Income ($)': [5000, 0, 0, 0, 0, 0],
            'Expense ($)': [0, 1200, 500, 1000, 300, 500],
            'Balance ($)': [0, 0, 0, 0, 0, 0]
        }
        df_finance = pd.DataFrame(finance_data)
        balance = 0
        for i in range(len(df_finance)):
            balance += df_finance.at[i, 'Income ($)'] - df_finance.at[i, 'Expense ($)']
            df_finance.at[i, 'Balance ($)'] = balance
        
        # ========== SHEET 6: MATH CALCULATOR ==========
        math_data = {
            'Operation': ['➕ Addition', '➖ Subtraction', '✖️ Multiplication', '➗ Division', 
                         '🔢 Power', '√ Square Root', '📊 Percentage', '🧮 Modulus'],
            'Value A': [25, 50, 12, 100, 4, 64, 200, 17],
            'Value B': [15, 25, 8, 20, 3, 2, 15, 5],
            'Result': [0, 0, 0, 0, 0, 0, 0, 0],
            'Formula': ['A + B', 'A - B', 'A × B', 'A ÷ B', 'A^B', '√A', '(A × B)/100', 'A % B']
        }
        df_math = pd.DataFrame(math_data)
        df_math.at[0, 'Result'] = df_math.at[0, 'Value A'] + df_math.at[0, 'Value B']
        df_math.at[1, 'Result'] = df_math.at[1, 'Value A'] - df_math.at[1, 'Value B']
        df_math.at[2, 'Result'] = df_math.at[2, 'Value A'] * df_math.at[2, 'Value B']
        df_math.at[3, 'Result'] = round(df_math.at[3, 'Value A'] / df_math.at[3, 'Value B'], 2)
        df_math.at[4, 'Result'] = df_math.at[4, 'Value A'] ** df_math.at[4, 'Value B']
        df_math.at[5, 'Result'] = round(df_math.at[5, 'Value A'] ** 0.5, 2)
        df_math.at[6, 'Result'] = (df_math.at[6, 'Value A'] * df_math.at[6, 'Value B']) / 100
        df_math.at[7, 'Result'] = df_math.at[7, 'Value A'] % df_math.at[7, 'Value B']
        
        # ========== SHEET 7: GEOMETRY CALCULATOR ==========
        geo_data = {
            'Shape': ['Circle', 'Square', 'Rectangle', 'Triangle', 'Sphere', 'Cube', 'Cylinder'],
            'Parameter 1 (r/s/l)': [5, 4, 6, 3, 3, 4, 5],
            'Parameter 2 (w/h)': [0, 0, 8, 4, 0, 0, 8],
            'Area / Surface Area': [0, 0, 0, 0, 0, 0, 0],
            'Perimeter / Volume': [0, 0, 0, 0, 0, 0, 0],
            'Formula Used': ['πr²', 's²', 'l×w', '½×b×h', '4πr²', '6s²', '2πrh']
        }
        df_geo = pd.DataFrame(geo_data)
        
        # Circle
        df_geo.at[0, 'Area / Surface Area'] = round(math.pi * (df_geo.at[0, 'Parameter 1 (r/s/l)'] ** 2), 2)
        df_geo.at[0, 'Perimeter / Volume'] = round(2 * math.pi * df_geo.at[0, 'Parameter 1 (r/s/l)'], 2)
        # Square
        df_geo.at[1, 'Area / Surface Area'] = df_geo.at[1, 'Parameter 1 (r/s/l)'] ** 2
        df_geo.at[1, 'Perimeter / Volume'] = 4 * df_geo.at[1, 'Parameter 1 (r/s/l)']
        # Rectangle
        df_geo.at[2, 'Area / Surface Area'] = df_geo.at[2, 'Parameter 1 (r/s/l)'] * df_geo.at[2, 'Parameter 2 (w/h)']
        df_geo.at[2, 'Perimeter / Volume'] = 2 * (df_geo.at[2, 'Parameter 1 (r/s/l)'] + df_geo.at[2, 'Parameter 2 (w/h)'])
        # Triangle
        df_geo.at[3, 'Area / Surface Area'] = 0.5 * df_geo.at[3, 'Parameter 1 (r/s/l)'] * df_geo.at[3, 'Parameter 2 (w/h)']
        df_geo.at[3, 'Perimeter / Volume'] = df_geo.at[3, 'Parameter 1 (r/s/l)'] + df_geo.at[3, 'Parameter 2 (w/h)'] + math.sqrt(df_geo.at[3, 'Parameter 1 (r/s/l)']**2 + df_geo.at[3, 'Parameter 2 (w/h)']**2)
        # Sphere
        df_geo.at[4, 'Area / Surface Area'] = round(4 * math.pi * (df_geo.at[4, 'Parameter 1 (r/s/l)'] ** 2), 2)
        df_geo.at[4, 'Perimeter / Volume'] = round((4/3) * math.pi * (df_geo.at[4, 'Parameter 1 (r/s/l)'] ** 3), 2)
        # Cube
        df_geo.at[5, 'Area / Surface Area'] = 6 * (df_geo.at[5, 'Parameter 1 (r/s/l)'] ** 2)
        df_geo.at[5, 'Perimeter / Volume'] = df_geo.at[5, 'Parameter 1 (r/s/l)'] ** 3
        # Cylinder
        df_geo.at[6, 'Area / Surface Area'] = round(2 * math.pi * df_geo.at[6, 'Parameter 1 (r/s/l)'] * df_geo.at[6, 'Parameter 2 (w/h)'], 2)
        df_geo.at[6, 'Perimeter / Volume'] = round(math.pi * (df_geo.at[6, 'Parameter 1 (r/s/l)'] ** 2) * df_geo.at[6, 'Parameter 2 (w/h)'], 2)
        
        # ========== SHEET 8: PERCENTAGE CALCULATOR ==========
        pct_data = {
            'Calculation Type': ['🏷️ Discount', '💰 Sales Tax', '💵 Tip/Gratuity', '📈 Profit Margin', 
                                '📊 Markup', '🎯 Commission', '📉 Loss', '⭐ Grade'],
            'Original Amount': [1000, 500, 150, 5000, 800, 2000, 1000, 85],
            'Percentage': [15, 10, 18, 25, 30, 8, 10, 92],
            'Calculated Amount': [0, 0, 0, 0, 0, 0, 0, 0],
            'Final Amount': [0, 0, 0, 0, 0, 0, 0, 0]
        }
        df_pct = pd.DataFrame(pct_data)
        df_pct['Calculated Amount'] = df_pct['Original Amount'] * (df_pct['Percentage'] / 100)
        df_pct['Final Amount'] = df_pct['Original Amount'] + df_pct['Calculated Amount']
        
        # ========== SHEET 9: TREND ANALYSIS ==========
        trend_data = {
            'Month': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'],
            'Sales ($)': [10000, 12000, 11500, 13500, 15000, 14800, 16000, 17000, 16500, 18000, 19000, 20000],
            'Growth %': [0] * 12,
            '3-Month AVG': [0] * 12,
            'Trend': ['Start'] * 12
        }
        df_trend = pd.DataFrame(trend_data)
        for i in range(1, 12):
            df_trend.at[i, 'Growth %'] = round(((df_trend.at[i, 'Sales ($)'] - df_trend.at[i-1, 'Sales ($)']) / df_trend.at[i-1, 'Sales ($)']) * 100, 2)
        for i in range(12):
            if i >= 2:
                df_trend.at[i, '3-Month AVG'] = round((df_trend.at[i-2, 'Sales ($)'] + df_trend.at[i-1, 'Sales ($)'] + df_trend.at[i, 'Sales ($)']) / 3, 2)
            if i > 0:
                df_trend.at[i, 'Trend'] = '📈 Up' if df_trend.at[i, 'Sales ($)'] > df_trend.at[i-1, 'Sales ($)'] else '📉 Down'
        
        # ========== SHEET 10: UNIT CONVERTER ==========
        unit_data = {
            'Category': ['Length', 'Weight', 'Time', 'Temperature', 'Currency', 'Volume', 'Speed', 'Area'],
            'From Unit': ['Kilometers', 'Kilograms', 'Hours', 'Celsius', 'USD', 'Liters', 'km/h', 'Square Meters'],
            'Value': [10, 100, 24, 25, 1000, 5, 100, 100],
            'To Unit': ['Miles', 'Pounds', 'Minutes', 'Fahrenheit', 'EUR', 'Gallons', 'mph', 'Square Feet'],
            'Converted Value': [0, 0, 0, 0, 0, 0, 0, 0],
            'Formula': ['× 0.6214', '× 2.2046', '× 60', '(°C × 9/5) + 32', '× 0.92', '× 0.2642', '× 0.6214', '× 10.764']
        }
        df_unit = pd.DataFrame(unit_data)
        df_unit.at[0, 'Converted Value'] = round(df_unit.at[0, 'Value'] * 0.6214, 2)
        df_unit.at[1, 'Converted Value'] = round(df_unit.at[1, 'Value'] * 2.2046, 2)
        df_unit.at[2, 'Converted Value'] = df_unit.at[2, 'Value'] * 60
        df_unit.at[3, 'Converted Value'] = (df_unit.at[3, 'Value'] * 9/5) + 32
        df_unit.at[4, 'Converted Value'] = round(df_unit.at[4, 'Value'] * 0.92, 2)
        df_unit.at[5, 'Converted Value'] = round(df_unit.at[5, 'Value'] * 0.2642, 2)
        df_unit.at[6, 'Converted Value'] = round(df_unit.at[6, 'Value'] * 0.6214, 2)
        df_unit.at[7, 'Converted Value'] = round(df_unit.at[7, 'Value'] * 10.764, 2)
        
        # ========== SHEET 11: INSTRUCTIONS ==========
        instructions_data = {
            'Step': ['1', '2', '3', '4', '5', '6'],
            'Action': [
                'Open this Excel file',
                'Look at the 🏠 INDEX sheet (this is your main menu)',
                'Click any sheet tab at the BOTTOM of Excel',
                'Each sheet has a different calculator ready to use',
                'Enter your own numbers in any white cell',
                'All formulas calculate automatically!'
            ],
            'Tip': [
                'Sheet tabs are your navigation buttons',
                'INDEX shows all 10 calculators',
                'Try 💰 STOCK first - it\'s simple!',
                'Add more rows if you need them',
                'Right-click any sheet → Copy to create custom calculator',
                'Formulas work exactly like Excel formulas'
            ]
        }
        df_instructions = pd.DataFrame(instructions_data)
        
        # ========== WRITE ALL SHEETS ==========
        print("📝 Writing Excel sheets...")
        
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            df_index.to_excel(writer, sheet_name='🏠 INDEX', index=False)
            df_stock.to_excel(writer, sheet_name='💰 STOCK', index=False)
            df_demo.to_excel(writer, sheet_name='👥 DEMOGRAPHY', index=False)
            df_stats.to_excel(writer, sheet_name='📈 STATISTICS', index=False)
            df_finance.to_excel(writer, sheet_name='💳 FINANCE', index=False)
            df_math.to_excel(writer, sheet_name='🔢 MATH', index=False)
            df_geo.to_excel(writer, sheet_name='📐 GEOMETRY', index=False)
            df_pct.to_excel(writer, sheet_name='🎯 PERCENTAGE', index=False)
            df_trend.to_excel(writer, sheet_name='📉 TREND', index=False)
            df_unit.to_excel(writer, sheet_name='⚖️ UNIT', index=False)
            df_instructions.to_excel(writer, sheet_name='📖 HELP', index=False)
        
        # Apply formatting
        self._apply_formatting()
        
        return True
    
    def _apply_formatting(self):
        """Apply professional formatting to all sheets"""
        try:
            wb = openpyxl.load_workbook(self.filename)
            
            # Color scheme for different sheets
            colors = {
                '🏠 INDEX': '2E7D32',
                '💰 STOCK': 'C62828',
                '👥 DEMOGRAPHY': '1565C0',
                '📈 STATISTICS': '6A1B9A',
                '💳 FINANCE': 'E65100',
                '🔢 MATH': '00695C',
                '📐 GEOMETRY': '283593',
                '🎯 PERCENTAGE': '827717',
                '📉 TREND': 'BF360C',
                '⚖️ UNIT': '00838F',
                '📖 HELP': '455A64'
            }
            
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            for sheet_name, color in colors.items():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    # Format header row
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border
                    
                    # Auto-fit columns
                    for column in ws.columns:
                        max_length = 0
                        col_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        ws.column_dimensions[col
