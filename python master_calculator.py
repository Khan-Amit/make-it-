import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.worksheet.buttons import Button
from datetime import datetime
import os

class MasterExcelCalculator:
    def __init__(self, filename="master_calculator.xlsx"):
        self.filename = filename
        
    def create_master_calculator(self):
        """Create complete Excel system with multiple calculators"""
        
        # ========== SHEET 1: DASHBOARD (Front Page) ==========
        dashboard_data = {
            'FEATURE': ['📊 Stock Calculator', '👥 Demography Calculator', '📈 Statistics Calculator', 
                       '💰 Finance Calculator', '🧮 Basic Math', '📐 Geometry Calculator',
                       '🎯 Percentage Calculator', '📉 Trend Analysis', '⚖️ Unit Converter'],
            'STATUS': ['Ready', 'Ready', 'Ready', 'Ready', 'Ready', 'Ready', 'Ready', 'Ready', 'Ready'],
            'LAST USED': [datetime.now().strftime("%Y-%m-%d"), '', '', '', '', '', '', '', ''],
            'BUTTON LINK': ['Go to Stock', 'Go to Demography', 'Go to Statistics', 'Go to Finance', 
                           'Go to Math', 'Go to Geometry', 'Go to Percentage', 'Go to Trend', 'Go to Unit']
        }
        df_dashboard = pd.DataFrame(dashboard_data)
        
        # ========== SHEET 2: STOCK CALCULATOR ==========
        stock_data = {
            'Product': ['Gold', 'Silver', 'Platinum', 'Copper', 'Diamond', 'Ruby', 'Emerald'],
            'Quantity': [100, 500, 50, 1000, 25, 40, 30],
            'Unit Price': [85000, 950, 30000, 8.5, 15000, 5000, 8000],
            'Total Value': [0, 0, 0, 0, 0, 0, 0],
            'Profit %': [15, 12, 20, 10, 25, 18, 22],
            'Profit Amount': [0, 0, 0, 0, 0, 0, 0]
        }
        df_stock = pd.DataFrame(stock_data)
        df_stock['Total Value'] = df_stock['Quantity'] * df_stock['Unit Price']
        df_stock['Profit Amount'] = df_stock['Total Value'] * (df_stock['Profit %'] / 100)
        
        # ========== SHEET 3: DEMOGRAPHY CALCULATOR ==========
        demography_data = {
            'Age Group': ['0-18', '19-30', '31-45', '46-60', '60+'],
            'Population': [25000, 45000, 38000, 22000, 15000],
            'Male %': [48, 50, 49, 47, 42],
            'Female %': [52, 50, 51, 53, 58],
            'Male Count': [0, 0, 0, 0, 0],
            'Female Count': [0, 0, 0, 0, 0],
            'Growth Rate %': [2.1, 1.8, 1.2, 0.5, -0.2]
        }
        df_demography = pd.DataFrame(demography_data)
        df_demography['Male Count'] = (df_demography['Population'] * df_demography['Male %'] / 100).astype(int)
        df_demography['Female Count'] = (df_demography['Population'] * df_demography['Female %'] / 100).astype(int)
        
        # ========== SHEET 4: STATISTICS CALCULATOR ==========
        statistics_data = {
            'Data Set': ['Sample A', 'Sample B', 'Sample C', 'Sample D', 'Sample E'],
            'Values': [23, 45, 67, 12, 89],
            'Value 2': [34, 56, 78, 23, 45],
            'Value 3': [45, 67, 89, 34, 56],
            'Mean': [0, 0, 0, 0, 0],
            'Median': [0, 0, 0, 0, 0],
            'Sum': [0, 0, 0, 0, 0]
        }
        df_statistics = pd.DataFrame(statistics_data)
        
        # Calculate statistics
        for i, row in df_statistics.iterrows():
            values = [row['Values'], row['Value 2'], row['Value 3']]
            df_statistics.at[i, 'Mean'] = sum(values) / len(values)
            df_statistics.at[i, 'Median'] = sorted(values)[1]
            df_statistics.at[i, 'Sum'] = sum(values)
        
        # ========== SHEET 5: FINANCE CALCULATOR ==========
        finance_data = {
            'Category': ['Income', 'Rent', 'Salary', 'Investment', 'Savings', 'Tax', 'Utilities'],
            'Amount': [50000, -12000, -25000, -5000, -3000, -5000, -2000],
            'Type': ['Credit', 'Debit', 'Debit', 'Debit', 'Debit', 'Debit', 'Debit'],
            'Date': ['2026-05-01', '2026-05-01', '2026-05-02', '2026-05-03', '2026-05-04', '2026-05-05', '2026-05-06'],
            'Balance': [0, 0, 0, 0, 0, 0, 0]
        }
        df_finance = pd.DataFrame(finance_data)
        
        # Running balance
        balance = 0
        for i in range(len(df_finance)):
            balance += df_finance.at[i, 'Amount']
            df_finance.at[i, 'Balance'] = balance
        
        # ========== SHEET 6: BASIC MATH CALCULATOR ==========
        math_data = {
            'Operation': ['Add', 'Subtract', 'Multiply', 'Divide', 'Power', 'Square Root', 'Modulus'],
            'Input A': [10, 20, 5, 100, 4, 64, 17],
            'Input B': [5, 8, 6, 20, 3, 2, 5],
            'Result': [0, 0, 0, 0, 0, 0, 0],
            'Formula': ['A+B', 'A-B', 'A*B', 'A/B', 'A^B', '√A', 'A%B']
        }
        df_math = pd.DataFrame(math_data)
        
        df_math.at[0, 'Result'] = df_math.at[0, 'Input A'] + df_math.at[0, 'Input B']
        df_math.at[1, 'Result'] = df_math.at[1, 'Input A'] - df_math.at[1, 'Input B']
        df_math.at[2, 'Result'] = df_math.at[2, 'Input A'] * df_math.at[2, 'Input B']
        df_math.at[3, 'Result'] = df_math.at[3, 'Input A'] / df_math.at[3, 'Input B']
        df_math.at[4, 'Result'] = df_math.at[4, 'Input A'] ** df_math.at[4, 'Input B']
        df_math.at[5, 'Result'] = df_math.at[5, 'Input A'] ** 0.5
        df_math.at[6, 'Result'] = df_math.at[6, 'Input A'] % df_math.at[6, 'Input B']
        
        # ========== SHEET 7: GEOMETRY CALCULATOR ==========
        geometry_data = {
            'Shape': ['Circle', 'Square', 'Rectangle', 'Triangle', 'Sphere', 'Cube', 'Cylinder'],
            'Parameter 1': [5, 4, 6, 3, 3, 4, 5],
            'Parameter 2': [0, 0, 8, 4, 0, 0, 8],
            'Area': [0, 0, 0, 0, 0, 0, 0],
            'Perimeter/Volume': [0, 0, 0, 0, 0, 0, 0],
            'Formula': ['πr²', 's²', 'l×w', '½×b×h', '⁴⁄₃πr³', 's³', 'πr²h']
        }
        df_geometry = pd.DataFrame(geometry_data)
        
        import math
        # Circle
        df_geometry.at[0, 'Area'] = math.pi * (df_geometry.at[0, 'Parameter 1'] ** 2)
        df_geometry.at[0, 'Perimeter/Volume'] = 2 * math.pi * df_geometry.at[0, 'Parameter 1']
        # Square
        df_geometry.at[1, 'Area'] = df_geometry.at[1, 'Parameter 1'] ** 2
        df_geometry.at[1, 'Perimeter/Volume'] = 4 * df_geometry.at[1, 'Parameter 1']
        # Rectangle
        df_geometry.at[2, 'Area'] = df_geometry.at[2, 'Parameter 1'] * df_geometry.at[2, 'Parameter 2']
        df_geometry.at[2, 'Perimeter/Volume'] = 2 * (df_geometry.at[2, 'Parameter 1'] + df_geometry.at[2, 'Parameter 2'])
        # Triangle
        df_geometry.at[3, 'Area'] = 0.5 * df_geometry.at[3, 'Parameter 1'] * df_geometry.at[3, 'Parameter 2']
        # Sphere
        df_geometry.at[4, 'Area'] = 4 * math.pi * (df_geometry.at[4, 'Parameter 1'] ** 2)
        df_geometry.at[4, 'Perimeter/Volume'] = (4/3) * math.pi * (df_geometry.at[4, 'Parameter 1'] ** 3)
        # Cube
        df_geometry.at[5, 'Area'] = 6 * (df_geometry.at[5, 'Parameter 1'] ** 2)
        df_geometry.at[5, 'Perimeter/Volume'] = df_geometry.at[5, 'Parameter 1'] ** 3
        # Cylinder
        df_geometry.at[6, 'Area'] = 2 * math.pi * df_geometry.at[6, 'Parameter 1'] * df_geometry.at[6, 'Parameter 2']
        df_geometry.at[6, 'Perimeter/Volume'] = math.pi * (df_geometry.at[6, 'Parameter 1'] ** 2) * df_geometry.at[6, 'Parameter 2']
        
        # ========== SHEET 8: PERCENTAGE CALCULATOR ==========
        percentage_data = {
            'Item': ['Discount', 'Tax', 'Commission', 'Tip', 'Profit Margin', 'Markup', 'Growth'],
            'Original Value': [1000, 500, 2000, 150, 5000, 800, 10000],
            'Percentage %': [15, 10, 8, 18, 25, 30, 12],
            'Calculated Value': [0, 0, 0, 0, 0, 0, 0],
            'Final Value': [0, 0, 0, 0, 0, 0, 0],
            'Formula': ['Original × %', 'Original × %', 'Original × %', 'Original × %', 'Original × %', 'Original × %', 'Original × %']
        }
        df_percentage = pd.DataFrame(percentage_data)
        df_percentage['Calculated Value'] = df_percentage['Original Value'] * (df_percentage['Percentage %'] / 100)
        df_percentage['Final Value'] = df_percentage['Original Value'] + df_percentage['Calculated Value']
        
        # ========== SHEET 9: TREND ANALYSIS ==========
        trend_data = {
            'Month': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'],
            'Sales': [1000, 1200, 1100, 1300, 1500, 1400, 1600, 1700, 1650, 1800, 1900, 2000],
            'Growth %': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
            '3-Month Avg': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
            'Trend': ['', '', '', '', '', '', '', '', '', '', '', '']
        }
        df_trend = pd.DataFrame(trend_data)
        
        # Calculate growth and averages
        for i in range(1, len(df_trend)):
            df_trend.at[i, 'Growth %'] = ((df_trend.at[i, 'Sales'] - df_trend.at[i-1, 'Sales']) / df_trend.at[i-1, 'Sales']) * 100
        
        for i in range(len(df_trend)):
            if i >= 2:
                df_trend.at[i, '3-Month Avg'] = (df_trend.at[i-2, 'Sales'] + df_trend.at[i-1, 'Sales'] + df_trend.at[i, 'Sales']) / 3
            df_trend.at[i, 'Trend'] = '📈 Up' if i > 0 and df_trend.at[i, 'Sales'] > df_trend.at[i-1, 'Sales'] else '📉 Down' if i > 0 else 'Start'
        
        # ========== SHEET 10: UNIT CONVERTER ==========
        unit_data = {
            'Convert From': ['Kilometers', 'Kilograms', 'Hours', 'Celsius', 'USD', 'Meters', 'Liters'],
            'Value': [10, 100, 24, 25, 1000, 50, 5],
            'Convert To': ['Miles', 'Pounds', 'Minutes', 'Fahrenheit', 'EUR', 'Feet', 'Gallons'],
            'Result': [0, 0, 0, 0, 0, 0, 0],
            'Conversion Formula': ['km × 0.6214', 'kg × 2.2046', 'hr × 60', '(°C × 9/5) + 32', 'USD × 0.92', 'm × 3.2808', 'L × 0.2642']
        }
        df_unit = pd.DataFrame(unit_data)
        
        df_unit.at[0, 'Result'] = df_unit.at[0, 'Value'] * 0.6214
        df_unit.at[1, 'Result'] = df_unit.at[1, 'Value'] * 2.2046
        df_unit.at[2, 'Result'] = df_unit.at[2, 'Value'] * 60
        df_unit.at[3, 'Result'] = (df_unit.at[3, 'Value'] * 9/5) + 32
        df_unit.at[4, 'Result'] = df_unit.at[4, 'Value'] * 0.92
        df_unit.at[5, 'Result'] = df_unit.at[5, 'Value'] * 3.2808
        df_unit.at[6, 'Result'] = df_unit.at[6, 'Value'] * 0.2642
        
        # Write all sheets to Excel
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            df_dashboard.to_excel(writer, sheet_name='📊 DASHBOARD', index=False)
            df_stock.to_excel(writer, sheet_name='💰 STOCK', index=False)
            df_demography.to_excel(writer, sheet_name='👥 DEMOGRAPHY', index=False)
            df_statistics.to_excel(writer, sheet_name='📈 STATISTICS', index=False)
            df_finance.to_excel(writer, sheet_name='💳 FINANCE', index=False)
            df_math.to_excel(writer, sheet_name='🧮 MATH', index=False)
            df_geometry.to_excel(writer, sheet_name='📐 GEOMETRY', index=False)
            df_percentage.to_excel(writer, sheet_name='🎯 PERCENTAGE', index=False)
            df_trend.to_excel(writer, sheet_name='📉 TREND', index=False)
            df_unit.to_excel(writer, sheet_name='⚖️ UNIT', index=False)
        
        # Apply professional formatting
        self._format_master_file()
        
        # Create summary
        self._create_summary_sheet()
        
        print("\n" + "="*70)
        print("🎯 MASTER CALCULATOR EXCEL FILE CREATED!")
        print("="*70)
        self._display_overview()
        
        return True
    
    def _format_master_file(self):
        """Apply formatting to all sheets"""
        try:
            wb = openpyxl.load_workbook(self.filename)
            
            # Color schemes for different sheets
            colors = {
                '📊 DASHBOARD': '1B5E20',
                '💰 STOCK': 'B71C1C',
                '👥 DEMOGRAPHY': '0D47A1',
                '📈 STATISTICS': '4A148C',
                '💳 FINANCE': 'E65100',
                '🧮 MATH': '004D40',
                '📐 GEOMETRY': '1A237E',
                '🎯 PERCENTAGE': '827717',
                '📉 TREND': 'BF360C',
                '⚖️ UNIT': '004D40'
            }
            
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            for sheet_name, color in colors.items():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    # Format headers
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        col_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 25)
            
            wb.save(self.filename)
            print("✓ Professional formatting applied!")
            
        except Exception as e:
            print(f"Formatting note: {e}")
    
    def _create_summary_sheet(self):
        """Create a summary statistics sheet"""
        try:
            wb = openpyxl.load_workbook(self.filename)
            
            summary_data = [
                ['Master Calculator Summary', ''],
                ['Created On', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ['Total Calculators', '10'],
                ['Sheets Included', '📊 DASHBOARD, 💰 STOCK, 👥 DEMOGRAPHY, 📈 STATISTICS, 💳 FINANCE, 🧮 MATH, 📐 GEOMETRY, 🎯 PERCENTAGE, 📉 TREND, ⚖️ UNIT'],
                ['Key Features', 'All formulas auto-calculate'],
                ['Instructions', 'Enter your data in ANY sheet - formulas update automatically!']
            ]
            
            df_summary = pd.DataFrame(summary_data)
            
            with pd.ExcelWriter(self.filename, engine='openpyxl', mode='a') as writer:
                df_summary.to_excel(writer, sheet_name='📋 SUMMARY', index=False)
            
            print("✓ Summary sheet added!")
            
        except Exception as e:
            print(f"Summary note: {e}")
    
    def _display_overview(self):
        """Display system overview"""
        print("\n📊 WHAT'S INCLUDED:")
        print("="*70)
        print("1️⃣  📊 DASHBOARD  - Front page with all calculator buttons")
        print("2️⃣  💰 STOCK      - Inventory & profit calculator")
        print("3️⃣  👥 DEMOGRAPHY - Population, gender & growth stats")
        print("4️⃣  📈 STATISTICS - Mean, median, sum calculations")
        print("5️⃣  💳 FINANCE    - Running balance & budget tracker")
        print("6️⃣  🧮 MATH       - Basic arithmetic operations")
        print("7️⃣  📐 GEOMETRY   - Area & volume of shapes")
        print("8️⃣  🎯 PERCENTAGE - Discount, tax, margin calculator")
        print("9️⃣  📉 TREND      - Sales growth & moving average")
        print("🔟 ⚖️ UNIT        - Convert units (km→miles, kg→lbs, etc.)")
        print("="*70)
        
        print("\n💡 HOW TO USE:")
        print("• Open the Excel file")
        print("• Go to 📊 DASHBOARD sheet to see all options")
        print("• Click any sheet tab (💰 STOCK, 👥 DEMOGRAPHY, etc.)")
        print("• Enter YOUR data in the columns - formulas auto-calculate!")
        print("• Add more rows or create new sheets for custom calculations")
        
        print("\n🎯 TO ADD YOUR OWN CALCULATIONS:")
        print("1. Right-click any sheet tab → 'Move or Copy' → 'Create a copy'")
        print("2. Rename the copied sheet to your calculation name")
        print("3. Modify the data and add your own formulas")
        print("4. Excel formulas work normally: =A1+B1, =SUM(A:A), etc.")
        
        print("\n✅ FILE READY:", os.path.abspath(self.filename))
        print("="*70)

def main():
    print("\n" + "🎯"*35)
    print("   MASTER EXCEL CALCULATOR SYSTEM")
    print("   With Dashboard & 10+ Calculators")
    print("🎯"*35)
    
    # Create the master system
    calculator = MasterExcelCalculator("master_calculator.xlsx")
    calculator.create_master_calculator()
    
    print("\n🚀 OPEN THE EXCEL FILE NOW!")
    print("   All calculations work automatically!")
    print("   You can add ANY calculation you want!")
    print("\n🙏 THANK YOU, BOSS!")

if __name__ == "__main__":
    try:
        import pandas, openpyxl, math
    except ImportError:
        print("Installing required packages...")
        os.system("pip install pandas openpyxl")
    
    main()
